'''
Name: AmazonWishListScrapper_c

Purpose: To scrap information off of Amazon wish lists.

Description: Given an Amazon Wish list in print view, this
    app will scrap information and save it to an excel spread sheet.

TO-DO:
-Create an excel file when one doesn't already exist.
-Open excel file after the script finishes running.
-Update to work on list that are not formatted for print view.
-Attach URL's to the items in the excel file.
-Attach to a script that runs automatically.
-Create a backup of the excel file
-Send an email when certain parameters have been reached.
    Such as a price has lowered under a specific threshold.


Created on Dec 18, 2017

@author: joshua
'''

import requests, sys, bs4, openpyxl, datetime

from openpyxl.utils  import get_column_letter
from openpyxl.styles import PatternFill

EXCEL_FILE_NAME = 'wishlist.xlsx'

DATE_TODAY = datetime.datetime.now().strftime("%m-%d-%Y")

#red is for the most expensive price
#green is for the cheapest price
#grey is for an item that is unavailable
#yellow is for an item that is no longer on the Amazon side
RED_FILL = PatternFill("solid", fgColor="FF6400")
GREEN_FILL = PatternFill("solid", fgColor="64FF00")
GREY_FILL = PatternFill("solid", fgColor="DDDDDD")
YELLOW_FILL = PatternFill("solid", fgColor="ffff00")
NO_FILL = PatternFill(fill_type=None)

mWriteItem = False
mMaxColumn = 1
mMaxRow = 1
ROW_START = 2
COLUMN_START = 2

mWishList = {}
                  
mWishListAddress = {'wishlist' : 'https://www.amazon.com/hz/wishlist/printview/PX0UCCNRH6TB',
                        'movies' : 'https://www.amazon.com/hz/wishlist/printview/2W5AXXPD3VYL9',
                        'comics' : 'https://www.amazon.com/hz/wishlist/printview/1R73OPLA6PM2H',
                        'toys' : 'https://www.amazon.com/hz/wishlist/printview/NUJRW2S5UINH',
                        'ComputerScience' : 'https://www.amazon.com/hz/wishlist/printview/1CIZE0Y8EV6WW',
                        'games' : 'https://www.amazon.com/hz/wishlist/printview/1E4ZJECB8UKAY',
                        'camera' : 'https://www.amazon.com/hz/wishlist/printview/1Q559HFNX3Q8I',
                        'dad' : 'https://www.amazon.com/hz/wishlist/printview/29PA5RXAV4IFY'
                         };

def run():
    """
    Run the main method of the program.
    """
    
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_NAME, read_only=False, guess_types=False);
    except IOError:
        print("IOError: " + input + " does not exist.")
        sys.exit(1)
    
    for title, addr in mWishListAddress.items():
        readExcel(workbook, title) 
        soup = dlWishList(addr, title)
        if(soup != None):
            readAmazon(soup)  
            writeExcel(workbook, title)  
            
        print()
        
    workbook.save('wishlist.xlsx')
    print('Done: Workbook saved')

def dlWishList(wishListAddr, title):
    """
    Download an Amazon Wishlist.

    Keyword arguments:
    wishListAddr -- the wishlist web address
    title -- the title of the Amazon wishlist
    
    Return:
    If successful, then return a beautiful soup object.
    Otherwise return none.  
    """
    print('Downloading: ', title, 'from', wishListAddr)
    
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.84 Safari/537.36',
    }
    try:
        url = wishListAddr
                
        res = requests.get(url, headers=header)
        res.raise_for_status()
        soup = bs4.BeautifulSoup(res.text, "html.parser")
        
        #return a soup object
        return soup;
    except requests.exceptions.HTTPError as e:
        print("Error: Could not connect to: " + wishListAddr)
        print(e)
        #return none
        return None;

def readAmazon(soup):
    """
    Read data in from from an Amazon Wishlist

    Keyword arguments:
    soup -- a Beautiful soup object to read from.
    """
    
    firstRow = True
    for row in soup.table.children:
        if(firstRow == False):
#             print(row.span.string)
            item = row.span.string
            cellCount = 0
            for cell in row:
                if(cellCount == 3):
#                     print(cell.string)
                    price = cell.string.strip()
                    price = price[1:]
                    try:
                        fPrice = float(price)
                    except:
                        fPrice = 00.00
    
                cellCount = cellCount + 1
                
            mWishList[item] = fPrice
        else:
            firstRow = False
                    
def readExcel(workbook, wbTitle):
    """
    Read from an excel file.

    Keyword arguments:
    workbook -- the workbook to write too
    wbTitle -- the workbook sheet to write to
    """
    print('Reading: ', wbTitle, ' from ', EXCEL_FILE_NAME)
    try:
        sheet = workbook.get_sheet_by_name(wbTitle)
    except:
        sheet = workbook.create_sheet(wbTitle)

    global mMaxColumn 
    mMaxColumn = sheet.max_column
    global mMaxRow
    mMaxRow = sheet.max_row
    
    if(mMaxColumn == 1):
        global mWriteItem 
        mWriteItem = True

    for row in range(ROW_START, mMaxRow):
        item = sheet['A' + str(row)].value
#         price = sheet[get_column_letter(mMaxColumn) + str(row)].value
        price = -1
        mWishList[item] = price

def writeExcel(workbook, wbTitle):
    """
    Write to an excel file.

    Keyword arguments:
    workbook -- the workbook to write too
    wbTitle -- the workbook sheet to write to
    """
    print('Writing: ', wbTitle, ' to ', EXCEL_FILE_NAME)
    
    sheet = workbook.get_sheet_by_name(wbTitle)

    columnToWriteTo = mMaxColumn
    dateMaxColumn = sheet[get_column_letter(mMaxColumn) + str(1)].value
    
    if(dateMaxColumn != None):
        if(DATE_TODAY != dateMaxColumn):
            #write to the next column
            columnToWriteTo = columnToWriteTo + 1
            sheet[get_column_letter(columnToWriteTo) + str(1)] = datetime.datetime.now().strftime("%m-%d-%Y")
        #else if the date is the same overwrite the column
    else:
        #if it's the first time writing to the file
        columnToWriteTo = columnToWriteTo + 1    
        sheet[get_column_letter(columnToWriteTo) + str(1)] = datetime.datetime.now().strftime("%m-%d-%Y")
    
    rowCount = ROW_START
    for item, price in mWishList.items():
        if(mWriteItem or sheet['A' + str(rowCount)].value == None):
            sheet['A' + str(rowCount)] = item
            
        sheet[get_column_letter(columnToWriteTo) + str(rowCount)] = price
        colorHighLow(sheet, rowCount, columnToWriteTo)
        rowCount = rowCount + 1
        
    mWishList.clear()

def colorHighLow(sheet, row, maxColumn):
    """
    Appy colors to significant cells.

    Keyword arguments:
    sheet -- the sheet being read/written too
    row -- the row being read/written too
    maxColumn -- the last column to read from
    """
    #Test for None in the first cell
    if (sheet[get_column_letter(COLUMN_START) + str(row)].value == None):
        sheet[get_column_letter(COLUMN_START) + str(row)].value = 0.0
        high = 0.0
        low = 0.0
    else:
        high = sheet[get_column_letter(COLUMN_START) + str(row)].value
        low =  sheet[get_column_letter(COLUMN_START) + str(row)].value
        sheet[get_column_letter(COLUMN_START) + str(row)].fill = NO_FILL
    
    #iterate through the row
    highCell = 'B' + str(row)
    lowCell = 'B' + str(row)
    for cell in range((COLUMN_START + 1), (maxColumn+1)):
        if(sheet[get_column_letter(cell) + str(row)].value == None):
            sheet[get_column_letter(cell) + str(row)].value = 0.0
            temp = 0.0
        else:
            temp = float(sheet[get_column_letter(cell) + str(row)].value)
            sheet[get_column_letter(cell) + str(row)].fill = NO_FILL
            
        if(temp >= high):
            high = temp
            highCell = get_column_letter(cell) + str(row)
        if(temp <= low or low == 0.0):
            low = temp
            lowCell = get_column_letter(cell) + str(row)
    
    #color the cells
    if(high != low):
        sheet[highCell].fill = RED_FILL
    if(low == -1.0):
        sheet[lowCell].fill = YELLOW_FILL
    if(low == 0.0):
        sheet[lowCell].fill = GREY_FILL
    if(low > 0.0):
        sheet[lowCell].fill = GREEN_FILL

def printWishList():
    for item, price in mWishList.items():
        print(item, price)  

if __name__ == '__main__':
    run()



    
    